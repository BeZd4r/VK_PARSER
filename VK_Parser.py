from openpyxl import load_workbook
import requests
import time

token = open("Server_Token_(VK_API).txt").read()
users = {}
offset = 0
count = 0
mem = {"1":"Подписчик","0":"Заблудший"}

params = {
        "access_token": f"{token}",
        "v":'5.131',
        "owner_id": "-212637974", 
        "count" : "100",
        "extended" : 1
    }

#create table with options
def Creat_Table(): 
    exel_file = "Tables.xlsx"

    work_book = load_workbook(exel_file)

    del work_book['Data']

    work_list = work_book.create_sheet("Data")
    
    work_list.append(["Номер", "Имя", "id", "Ссылка", "Лайки", "Комментарии"])

    work_list.column_dimensions['A'].width = 10
    work_list.column_dimensions['B'].width = 24
    work_list.column_dimensions['C'].width = 15
    work_list.column_dimensions['D'].width = 30
    work_list.column_dimensions['E'].width = 15
    work_list.column_dimensions['F'].width = 20
    work_list.column_dimensions['G'].width = 20


    i = 1
    for obj in users:
        res = [str(users[obj][subobj]) for subobj in users[obj].keys()[:-1]]
        res.insert(0,i)
        work_list.append(res)
        
        i += 1

    work_book.save(exel_file)
    work_book.close()


def Post_Checker(post_id):
    
    def Check_comm(response):
        print(response)
        response = response["profiles"] + response["groups"]
        for objects in response:

            if str(objects['id']) in users.keys():

                users[str(objects["id"])]["comments"] += 1

            else:

                if "name" not in objects.keys():
                    users[str(objects["id"])] = {
                        "name": f"{objects['first_name']} {objects['last_name']}",
                        "id" : str(objects['id']),
                        "link" : "https://vk.com/id"+str(objects['id']),
                        "likes" : 0,
                        "comments" : 1
                        }
                else:
                    users[str(objects["id"])] = {
                        "name": f"{objects['name']}",
                        "id" : str(objects['id']),
                        "link" : "https://vk.com/id"+str(objects['id']),
                        "likes" : 0,
                        "comments" : 1
                        }
                


    def Check_likes(response):
        response = response["items"]
        print(response)
        for objects in response:

            if str(objects['id']) in users.keys():

                users[str(objects["id"])]["likes"] += 1

            else:

                if "name" not in objects.keys():
                    users[str(objects["id"])] = {
                        "name": f"{objects['first_name']} {objects['last_name']}",
                        "id" : str(objects['id']),
                        "link" : "https://vk.com/id"+str(objects['id']),
                        "likes" : 1,
                        "comments" : 0
                        }
                else:
                    users[str(objects["id"])] = {
                        "name": f"{objects['name']}",
                        "id" : str(objects['id']),
                        "link" : "https://vk.com/id"+str(objects['id']),
                        "likes" : 1,
                        "comments" : 0
                        }
        
    # try:

    url_post_comm = f"https://api.vk.com/method/wall.getComments?thread_items_count=10&post_id={post_id}"
    url_post_likes = f"https://api.vk.com/method/likes.getList?item_id={post_id}&type=post"

    Check_comm(requests.get(url_post_comm,params).json()["response"])
    # time.sleep(0.25)
    Check_likes(requests.get(url_post_likes,params).json()['response'])

    print(f"Post {post_id} is checked")

    # except Exception:
    #     print(f"Post {post_id} is broke")
        
def Wall_Checker():
    global offset, count
    url_wall = f"https://api.vk.com/method/wall.get?offset={offset}&count=100"
    response = requests.get(url_wall,params).json()
    

    for item in response['response']['items']:
        count += 1
        print(count)
        Post_Checker(int(item['id']))
        time.sleep(0.25)
    offset += 100

    if offset > response['response']['count']:
        return
    else:     
     Wall_Checker()


Wall_Checker()

print(users)

Creat_Table()

    